#!/usr/bin/env python
# coding=utf-8
import traceback

import openpyxl as oxl
from PyQt5.QtCore import QThread, pyqtSignal

import myUtils


class ExportExcellThread(QThread):
    signal_end = pyqtSignal(bool, str)
    signal_log = pyqtSignal(str)

    def __init__(self, urlTable, sensiveResultTable,otherDomainResultTable, saveCount=1000):
        super(ExportExcellThread, self).__init__()
        self.urlTable = urlTable
        self.sensiveResultTable = sensiveResultTable
        self.otherDomainResultTable = otherDomainResultTable
        self.saveCount = saveCount

    def run(self):
        filename = "导出文件-" + myUtils.getNowSeconed().replace("-", "").replace(" ", "").replace(":", "") + ".xlsx"
        resultFlag = False
        logStr = ""
        self.signal_log.emit("导出文件名为：{}".format(filename))
        try:
            nowUrlResultCount = self.urlTable.rowCount()
            nowSensiveResultCount = self.sensiveResultTable.rowCount()
            nowOtherDomainResultCount = self.otherDomainResultTable.rowCount()
            # 创建一个excell文件对象
            wb = oxl.Workbook()
            # 创建URL扫描结果子表
            ws = wb.active
            ws.title = "URL扫描结果"
            # 创建表头
            headArr = ["序号", "URL", "响应码", "URL类型", "标题", "包长度", "状态"]
            myUtils.writeExcellHead(ws, headArr)

            # 遍历当前结果
            self.signal_log.emit("开始导出URL扫描结果")
            for rowIndex in range(nowUrlResultCount):
                if rowIndex % self.saveCount == 0:
                    minIndex = rowIndex + 1
                    maxIndex = rowIndex + self.saveCount if nowUrlResultCount > rowIndex + self.saveCount else nowUrlResultCount
                    tmpLogStr = "正在导出{0}-{1}行数据".format(minIndex, maxIndex)
                    self.signal_log.emit(tmpLogStr)
                else:
                    pass
                # 获取当前行的值
                nowUrl = self.urlTable.item(rowIndex, 0).text()
                nowStatus = self.urlTable.item(rowIndex, 1).text()
                nowTitle = self.urlTable.item(rowIndex, 2).text()
                nowContentLength = self.urlTable.item(rowIndex, 3).text()
                nowUrlType = self.urlTable.item(rowIndex, 4).text()
                nowState = "被过滤" if int(self.urlTable.item(rowIndex, 5).text()) == 0 else "正常"

                # 将值写入excell对象
                myUtils.writeExcellCell(ws, rowIndex + 2, 1, rowIndex + 1, 0, True)
                myUtils.writeExcellCell(ws, rowIndex + 2, 2, nowUrl, 0, False, hyperLink=nowUrl)
                myUtils.writeExcellCell(ws, rowIndex + 2, 3, nowStatus, 0, True)
                myUtils.writeExcellCell(ws, rowIndex + 2, 4, nowUrlType, 0, True)
                myUtils.writeExcellCell(ws, rowIndex + 2, 5, nowTitle, 0, False)
                myUtils.writeExcellCell(ws, rowIndex + 2, 6, nowContentLength, 0, False)
                myUtils.writeExcellCell(ws, rowIndex + 2, 7, nowState, 0, True)
                myUtils.writeExcellSpaceCell(ws, rowIndex + 2, 8)

                # 指定数量行保存一次
                if rowIndex != 0 and rowIndex % self.saveCount == 0:
                    myUtils.saveExcell(wb, saveName=filename)
                    wb = oxl.load_workbook(filename)
                    ws = wb.get_sheet_by_name(wb.get_sheet_names()[0])

            # 设置列宽
            colWidthArr = [7, 70, 7, 10, 60, 10, 10]
            myUtils.setExcellColWidth(ws, colWidthArr)

            myUtils.saveExcell(wb, saveName=filename)
            wb = oxl.load_workbook(filename)
            self.signal_log.emit("URL扫描结果导出成功")

            # 创建敏感信息扫描结果子表
            self.signal_log.emit("开始导出敏感信息扫描结果")
            ws = wb.create_sheet("敏感信息扫描结果", 1)
            # 创建表头
            headArr = ["序号", "URL", "关键词", "敏感信息"]
            myUtils.writeExcellHead(ws, headArr)

            # 遍历当前结果
            for rowIndex in range(nowSensiveResultCount):
                if rowIndex % self.saveCount == 0:
                    minIndex = rowIndex + 1
                    maxIndex = rowIndex + self.saveCount if nowSensiveResultCount > rowIndex + self.saveCount else nowSensiveResultCount
                    tmpLogStr = "正在导出{0}-{1}行数据".format(minIndex, maxIndex)
                    self.signal_log.emit(tmpLogStr)
                else:
                    pass
                # 获取当前行的值
                nowUrl = self.sensiveResultTable.item(rowIndex, 0).text()
                nowKey = self.sensiveResultTable.item(rowIndex, 1).text()
                nowSensiveVal = self.sensiveResultTable.item(rowIndex, 2).text()

                # 将值写入excell对象
                myUtils.writeExcellCell(ws, rowIndex + 2, 1, rowIndex + 1, 0, True)
                myUtils.writeExcellCell(ws, rowIndex + 2, 2, nowUrl, 0, False, hyperLink=nowUrl)
                myUtils.writeExcellCell(ws, rowIndex + 2, 3, nowKey, 0, True)
                myUtils.writeExcellCell(ws, rowIndex + 2, 4, nowSensiveVal, 0, False)
                myUtils.writeExcellSpaceCell(ws, rowIndex + 2, 5)
                # 指定数量行保存一次
                if rowIndex != 0 and rowIndex % self.saveCount == 0:
                    myUtils.saveExcell(wb, saveName=filename)
                    wb = oxl.load_workbook(filename)
                    ws = wb.get_sheet_by_name(wb.get_sheet_names()[1])

            # 设置列宽
            colWidthArr = [7, 70, 7, 60]
            myUtils.setExcellColWidth(ws, colWidthArr)

            # 创建其他域名扫描结果子表
            self.signal_log.emit("开始导出其他域名扫描结果")
            ws = wb.create_sheet("其他域名扫描结果", 2)
            # 创建表头
            headArr = ["序号", "域名", "URL"]
            myUtils.writeExcellHead(ws, headArr)

            # 遍历当前结果
            for rowIndex in range(nowOtherDomainResultCount):
                if rowIndex % self.saveCount == 0:
                    minIndex = rowIndex + 1
                    maxIndex = rowIndex + self.saveCount if nowOtherDomainResultCount > rowIndex + self.saveCount else nowOtherDomainResultCount
                    tmpLogStr = "正在导出{0}-{1}行数据".format(minIndex, maxIndex)
                    self.signal_log.emit(tmpLogStr)
                else:
                    pass
                # 获取当前行的值
                nowDomain = self.otherDomainResultTable.item(rowIndex, 0).text()
                nowUrl = self.otherDomainResultTable.item(rowIndex, 1).text()

                # 将值写入excell对象
                myUtils.writeExcellCell(ws, rowIndex + 2, 1, rowIndex + 1, 0, True)
                myUtils.writeExcellCell(ws, rowIndex + 2, 2, nowDomain, 0, False)
                myUtils.writeExcellCell(ws, rowIndex + 2, 3, nowUrl, 0, False,hyperLink=nowUrl)
                myUtils.writeExcellSpaceCell(ws, rowIndex + 2, 4)
                # 指定数量行保存一次
                if rowIndex != 0 and rowIndex % self.saveCount == 0:
                    myUtils.saveExcell(wb, saveName=filename)
                    wb = oxl.load_workbook(filename)
                    ws = wb.get_sheet_by_name(wb.get_sheet_names()[1])

            # 设置列宽
            colWidthArr = [7, 60, 70]
            myUtils.setExcellColWidth(ws, colWidthArr)

            # 保存文件
            myUtils.saveExcell(wb, saveName=filename)
            self.signal_log.emit("URL扫描结果导出成功")
            resultFlag = True
            logStr = "成功保存文件：{0}.xlsx 至当前文件夹".format(filename)
            self.signal_end.emit(resultFlag, logStr)
        except Exception as ex:
            # 删除可能已经生成的保存文件
            # if os.path.exists(filename):
            #     self.signal_log.emit("导出发生异常，回滚文件：{0}".format(filename))
            #     os.remove(filename)
            # else:
            #     pass
            resultFlag = False
            logStr = "保存文件失败，报错信息为：{0}".format(traceback.format_exc())
            self.signal_end.emit(resultFlag, logStr)
