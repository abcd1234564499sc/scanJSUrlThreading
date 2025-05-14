#!/usr/bin/env python
# coding=utf-8
import datetime
import hashlib
import json
import os
import random
import re
import socket
import urllib.parse
import warnings

import openpyxl as oxl
import requests
from openpyxl.styles import Border, Side, Font, PatternFill

# 全局变量区域
borderNumDic = {-1: None, 0: "thin"}

# 获得一个用于模拟浏览器的header
def getBrowerHeader():
    header = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/75.0.3770.100 Safari/537.36"
    }
    return header

def findHtmlTitleWithContent(htmlContent):
    title = ""
    reTitleList = re.findall(r"<title.*?>(.+?)</title>", htmlContent)
    title = "None" if len(reTitleList) == 0 else reTitleList[0]
    return title

# 访问URL,type表示请求类型，0为GET，1为POST，2为PUT
# 返回值类型如下：
# {
# "url":传入URL,
# "resultStr":访问结果字符串,
# "checkFlag":标志是否访问成功的布尔类型变量,
# "pageContent":访问成功时的页面源码，
# "status":访问的响应码,
# "responseHeaders":响应头，是一个字典,
# "responseCostSecond":响应时间（秒）,
# "response":requests库请求后返回的response对象
# }
def requestsUrl(url, cookie={}, header={}, data={}, type=0, reqTimeout=10, readTimeout=10, proxies=None,
                dataType="form", files={}):
    warnings.filterwarnings("ignore")
    resDic = {}
    url = url.strip()

    resultStr = ""
    checkedFlag = False
    status = ""
    title = ""
    pageContent = ""
    reContent = ""
    responseHeaders = {}
    timeout = (reqTimeout, readTimeout)
    header = header if header != {} else getBrowerHeader()
    response = None
    responseCostSecond = 0

    try:
        if type == 0:
            response = requests.get(url, headers=header, verify=False, cookies=cookie, timeout=timeout,
                                    proxies=proxies)
        elif type == 1:
            if dataType == "form":
                response = requests.post(url, headers=header, verify=False, cookies=cookie, data=data, timeout=timeout,
                                         proxies=proxies)
            elif dataType == "json":
                response = requests.post(url, headers=header, verify=False, cookies=cookie, json=data, timeout=timeout,
                                         proxies=proxies)
            elif dataType == "files":
                response = requests.post(url, headers=header, verify=False, cookies=cookie, data=data, files=files,
                                         timeout=timeout,
                                         proxies=proxies)
            else:
                response = requests.post(url, headers=header, verify=False, cookies=cookie, data=data, timeout=timeout,
                                         proxies=proxies)
        elif type == 2:
            response = requests.put(url, headers=header, verify=False, cookies=cookie, data=data, timeout=timeout,
                                    proxies=proxies)
        else:
            pass
        status = response.status_code
        responseHeaders = dict(response.headers)
        responseCostSecond = response.elapsed.total_seconds()
        if str(status)[0] == "2" or str(status)[0] == "3":
            # 获得页面编码
            pageEncoding = response.apparent_encoding
            # 设置页面编码
            response.encoding = pageEncoding
            # 获得页面内容
            reContent = response.text
            title = findHtmlTitleWithContent(reContent)
            resultStr = "验证成功，标题为：{0}".format(title)
            checkedFlag = True
        elif str(status)[0] == "4":
            title = "404 Not Found"
            resultStr = "请求资源不存在"
            checkedFlag = True
        else:
            resultStr = "验证失败，状态码为{0}".format(status)
            checkedFlag = False
    except Exception as e:
        resultStr = str(e)
        checkedFlag = False

    # 构建返回结果
    resDic["url"] = url
    resDic["resultStr"] = resultStr
    resDic["checkFlag"] = checkedFlag
    resDic["title"] = title
    resDic["status"] = status
    resDic["pageContent"] = reContent
    resDic["responseHeaders"] = responseHeaders
    resDic["responseCostSecond"] = responseCostSecond
    resDic["response"] = response
    return resDic


# 判断该字符串是否是IP，返回一个布尔值
def ifIp(matchStr):
    reFlag = True
    ipReg = r"^(?:(?:25[0-5]|2[0-4][0-9]|[01]?[0-9][0-9]?)\.){3}(?:25[0-5]|2[0-4][0-9]|[01]?[0-9][0-9]?)$"
    if re.match(ipReg, matchStr):
        reFlag = True
    else:
        reFlag = False
    return reFlag


# 从URL中提取协议+域名（端口）部分
def getUrlDomain(url):
    reDomainStr = ""
    urlObj = urllib.parse.urlsplit(url)
    tmpDomain = urlObj.netloc
    if tmpDomain != "":
        reDomainStr = urllib.parse.urlunsplit(tuple(list(urlObj[:2]) + [""] * 3))
    else:
        pass
    return reDomainStr

# 从URL中提取域名部分
def getUrlOnlyDomain(url):
    reDomainStr = ""
    urlObj = urllib.parse.urlsplit(url)
    tmpDomain = urlObj.netloc
    if tmpDomain != "":
        # 去除端口号
        tmpPortStartIndex = tmpDomain.find(":")
        if tmpPortStartIndex!=-1:
            reDomainStr = tmpDomain[:tmpPortStartIndex]
        else:
            reDomainStr = tmpDomain
    else:
        pass
    return reDomainStr

# 从URL中提取主域名部分（例：test1.test2.maindomain.com 的主域名为 maindomain.com）,IP会直接返回整个IP
def getUrlMainDomain(url):
    reDomainStr = ""
    tmpDomainStr = getUrlOnlyDomain(url)
    if tmpDomainStr.count(".") < 2:
        reDomainStr = tmpDomainStr
    else:
        reDomainStr = ".".join(tmpDomainStr.split(".")[-2:])
    return reDomainStr

# 从URL中提取协议+域名+path部分(不包含文件名或最后一个路径)
def getUrlWithoutLastPath(url):
    reUrlStr = ""
    urlObj = urllib.parse.urlsplit(url)
    tmpPath = urlObj.path
    tmpSolvedPath = "/".join(tmpPath.split("/")[:-1])

    reUrlStr = urllib.parse.urlunsplit(tuple(list(urlObj[:2])+[tmpSolvedPath] + [""] * 2))

    return reUrlStr

# 若URL以.xx结尾，则去除访问文件，只返回目录，否则直接返回
def getUrlWithoutFilePath(url):
    reUrlStr = ""
    urlObj = urllib.parse.urlsplit(url)
    tmpPath = urlObj.path
    tmpSolvedPath = "/".join(tmpPath.split("/")[:-1])
    tmpUrlFileName = tmpPath.split("/")[-1]
    if tmpUrlFileName.find(".")!=-1:
        reUrlStr = urllib.parse.urlunsplit(tuple(list(urlObj[:2]) + [tmpSolvedPath] + [""] * 2))
    else:
        reUrlStr = urllib.parse.urlunsplit(tuple(list(urlObj[:3])+ [""] * 2))

    return reUrlStr

# 返回URL的path及之后字符串
def getUrlStartWithPath(url):
    reUrlStr = ""
    urlObj = urllib.parse.urlsplit(url)
    reUrlStr = urllib.parse.urlunsplit(tuple([""]*2+list(urlObj[2:])))
    return reUrlStr


# 判断两个域名是否属于同一主域名，如果两者都是IP则比较是否完全相同，
# 返回一个布尔值,属于返回True,否则返回False
def ifSameMainDomain(domain1, domain2):
    domain1 = domain1.strip()
    domain2 = domain2.strip()

    if ifIp(domain1) and ifIp(domain2):
        # 两者都是IP，比较是否完全相同
        reFlag = (domain1 == domain2)
    elif (not ifIp(domain1)) and (not ifIp(domain2)):
        # 两个参数都是域名
        tmpMainDomainStr1 = getUrlMainDomain(domain1)
        tmpMainDomainStr2 = getUrlMainDomain(domain2)
        reFlag = (tmpMainDomainStr1 == tmpMainDomainStr2)
    else:
        # 两者类型不同，认为不同
        reFlag = False
    return reFlag

def splitUrlStr(urlStr,defaultIfHttpsBool=False):
    reDict = {}
    reDict["ifHttpsBool"] = False
    reDict["host"] = ""
    reDict["port"] = ""
    reDict["uri"] = ""
    reDict["queryStr"] = ""

    tmpUrlParseObj = urllib.parse.urlparse(urlStr)
    tmpIfHttpsBool = False
    if tmpUrlParseObj.scheme == "":
        tmpIfHttpsBool = defaultIfHttpsBool
    else:
        tmpIfHttpsBool = (tmpUrlParseObj.scheme == "https")
    reDict["ifHttpsBool"] = tmpIfHttpsBool
    reDict["uri"] = tmpUrlParseObj.path.lstrip("/")

    # 解析主机和端口
    tmpNetloc = tmpUrlParseObj.netloc
    ifGetPortFlag = False
    if ":" in tmpNetloc:
        tmpSplitList = tmpNetloc.split(":")
        if len(tmpSplitList) == 2:
            ifGetPortFlag = True
            reDict["host"] = tmpSplitList[0].strip()
            reDict["port"] = tmpSplitList[1].strip()
        else:
            pass
    else:
        pass

    if not ifGetPortFlag:
        # 根据协议设置默认端口
        reDict["host"] = tmpNetloc
        if reDict["ifHttpsBool"]:
            reDict["port"] = "443"
        else:
            reDict["port"] = "80"

    # 解析请求参数
    reDict["reqData"] = tmpUrlParseObj.query

    return reDict

# 判断两个URL的domain部分（域名/IP+端口）是否完全相同，
# 返回一个布尔值,相同返回True,否则返回False
def ifSameDomainWithTwoUrl(url1,url2):
    tmpUrl1SplitedDict = splitUrlStr(url1)
    tmpUrl2SplitedDict = splitUrlStr(url2)

    tmpCheckDomainStr1 = f"{'https' if tmpUrl1SplitedDict['ifHttpsBool'] else 'http'}://{tmpUrl1SplitedDict['host']}:{tmpUrl1SplitedDict['port']}"
    tmpCheckDomainStr2 = f"{'https' if tmpUrl2SplitedDict['ifHttpsBool'] else 'http'}://{tmpUrl2SplitedDict['host']}:{tmpUrl2SplitedDict['port']}"

    return (tmpCheckDomainStr1==tmpCheckDomainStr2)

# 从URL中提取文件后缀名
def getUrlFileSuffix(url):
    reSuffix = ""
    url = url.strip("/")


    tmpUrlParseResult = urllib.parse.urlparse(url)
    tmpPath = tmpUrlParseResult.path
    tmpUrlFilePath = tmpPath.split("/")[-1]
    if tmpUrlFilePath.find(".")!=-1:
        reSuffix = tmpUrlFilePath.split(".")[-1].strip().lower()
    else:
        reSuffix = ""

    return reSuffix


# 获得精确到秒的当前时间
def getNowSeconed():
    formatStr = "%Y-%m-%d %H:%M:%S"
    nowDate = datetime.datetime.now()
    nowDateStr = nowDate.strftime(formatStr)
    return nowDateStr


# 删除指定路径的文件,传入一个绝对路径,返回一个布尔变量以及一个字符串变量，
# 布尔变量为True表示是否删除成功,若为False则字符串变量中写入错误信息
def deleteFile(filePath):
    deleteFlag = True
    reStr = ""
    if os.path.exists(filePath):
        try:
            os.remove(filePath)
        except Exception as ex:
            reStr = "删除失败，失败信息为：{0}".format(ex)
            deleteFlag = False
    else:
        reStr = "未找到指定路径的文件"
        deleteFlag = False
    return deleteFlag, reStr


# 获得excell的常用样式
def getExcellStyleDic():
    styleDic = {}

    # 单线边框
    thinBorder = Border(left=Side(border_style='thin', color='000000'),
                        right=Side(border_style='thin', color='000000'),
                        top=Side(border_style='thin', color='000000'),
                        bottom=Side(border_style='thin', color='000000'))

    # 文字居中
    alignStyle = oxl.styles.Alignment(horizontal='center', vertical='center')
    leftStyle = oxl.styles.Alignment(horizontal='left', vertical='center')
    rightStyle = oxl.styles.Alignment(horizontal='right', vertical='center')

    # 加粗字体
    boldFont = Font(bold=True)
    hyperLinkFont = Font(color='0000FF')
    underLineFont = Font(underline='single')

    styleDic["thin"] = thinBorder
    styleDic["align"] = alignStyle
    styleDic["bold"] = boldFont
    styleDic["left"] = leftStyle
    styleDic["right"] = rightStyle
    styleDic["link"] = hyperLinkFont
    styleDic["underLine"] = underLineFont
    return styleDic


# 写入一个标准的excell表头（居中，单线框，加粗）
def writeExcellHead(ws, headArr):
    # 获得常用样式
    styleDic = getExcellStyleDic()
    # 写入表头
    for index, head in enumerate(headArr):
        ws.cell(row=1, column=index + 1).value = head
        ws.cell(row=1, column=index + 1).border = styleDic["thin"]
        ws.cell(row=1, column=index + 1).alignment = styleDic["align"]
        ws.cell(row=1, column=index + 1).font = styleDic["bold"]
    return ws


# 写入一个内容单元格
# borderNum表示该单元格的边框对象，其值可查询全局变量styleDic
# ifAlign是一个boolean对象，True表示居中
# hyperLink表示该单元格指向的链接，默认为None，表示不指向任何链接
# fgColor表示该单元格的背景颜色，为一个RGB16进制字符串，默认为“FFFFFF”（白色）
# otherAlign表示当ifAlign为False时指定的其他对齐方式，是一个数字型变量，默认为None，当其为0时表示左对齐，1为右对齐
def writeExcellCell(ws, row, column, value, borderNum, ifAlign, hyperLink=None, fgColor="FFFFFF", otherAlign=None):
    value = str(value)
    ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')
    value = ILLEGAL_CHARACTERS_RE.sub("", value)
    # 获得常用样式
    styleDic = getExcellStyleDic()
    # 获得指定单元格
    aimCell = ws.cell(row=row, column=column)
    # 设置值
    aimCell.value = value
    # 设置边框
    styleObjKey = borderNumDic[borderNum]
    if not styleObjKey:
        pass;
    else:
        styleObj = styleDic[styleObjKey]
        aimCell.border = styleObj
    # 设置居中
    if ifAlign:
        aimCell.alignment = styleDic["align"]
    elif otherAlign is not None:
        otherAlign = int(otherAlign)
        if otherAlign == 0:
            aimCell.alignment = styleDic["left"]
        else:
            aimCell.alignment = styleDic["right"]
    else:
        pass

    # 设置超链接
    if hyperLink:
        # 写入超链接
        aimCell.hyperlink = hyperLink
        # 设置当前单元格字体颜色为深蓝色，并添加下划线
        aimCell.font = styleDic["link"]
    else:
        pass

    # 设置填充颜色
    fill = PatternFill("solid", fgColor=fgColor)
    aimCell.fill = fill

    return ws


# 写入一个空格单元格，防止上一列文本超出
def writeExcellSpaceCell(ws, row, column):
    # 设置值
    ws.cell(row=row, column=column).value = " "

    return ws


# 设置excell的列宽
def setExcellColWidth(ws, colWidthArr):
    for colWidindex in range(len(colWidthArr)):
        ws.column_dimensions[chr(ord("A") + colWidindex)].width = colWidthArr[colWidindex]

    return ws


# 保存excell文件
def saveExcell(wb, saveName):
    savePath = ""
    # 处理传入的文件名
    saveName = saveName.split(".")[0] + ".xlsx"
    savePath = "{0}\\{1}".format(os.getcwd(), saveName)

    # 检测当前目录下是否有该文件，如果有则清除以前保存文件
    if os.path.exists(savePath):
        deleteFile(savePath)
    wb.save(savePath)
    return True


# 根据传入的contentDic写入配置文件，若配置文件不存在会创建
# contentDic格式为：{"配置名":"配置值"}，
# 对list和dict格式的值会使用json.dumps进行转换
def writeToConfFile(filePath, contentDic):
    with open(filePath, "w+", encoding="utf-8") as fr:
        for key, value in contentDic.items():
            if isinstance(value, list) or isinstance(value, dict):
                value = json.dumps(value)
            content = "{0}={1}".format(key, value)
            fr.write(content + "\n")


# 读取配置文件，生成一个配置字典，
# 字典结构为：{"配置名":"配置值"}，
# 文件结构为：配置名=配置值，对list和dict格式的值会使用json.loads进行转换
# 配置字典的最后一项固定为"confHeader":配置名列表
def readConfFile(filePath):
    confDic = {}
    headerList = []
    with open(filePath, "r", encoding="utf-8") as fr:
        fileLines = fr.readlines()
    fileLines = [a.replace("\r\n", "\n").replace("\n", "") for a in fileLines if a != ""]
    for fileLine in fileLines:
        fileLine = fileLine.replace("\r\n", "\n").replace("\n", "")
        tempList = fileLine.split("=")
        key = tempList[0].strip()
        value = "=".join(tempList[1:]).strip()
        try:
            value = json.loads(value)
        except:
            pass
        confDic[key] = value
        headerList.append(key)
    confDic["confHeader"] = headerList

    return confDic


def joinUrl(mainUrl, link):
    completeUrl = ""

    tmpMainUrlSplitedObj = urllib.parse.urlsplit(mainUrl)
    tmpMainUri = tmpMainUrlSplitedObj.path
    tmpMainUrlDomainStr = getUrlDomain(mainUrl)
    if tmpMainUri != "" and tmpMainUri[-1]!="/":
        tmpMainUri += "/"
    if link.startswith("/"):
        link = link[1:]

    tmpArgStr = ""
    tmpFindIndex = link.find("?")
    tmpUriStr = link
    if tmpFindIndex != -1:
        tmpArgStr = link[tmpFindIndex:]
        tmpUriStr = link[:tmpFindIndex]
    tmpUriStr = urllib.parse.urljoin("/",tmpMainUri+tmpUriStr)
    tmpUriStr += tmpArgStr
    if tmpMainUrlDomainStr != "" and tmpMainUrlDomainStr[-1] == "/":
        tmpMainUrlDomainStr = tmpMainUrlDomainStr[:-1]
    if tmpUriStr != "" and tmpUriStr[0]!="/":
        tmpUriStr = "/" + tmpUriStr
    completeUrl = tmpMainUrlDomainStr + tmpUriStr
    return completeUrl


# 测试指定IP的指定端口是否能联通
def connectIpPort(ip, port):
    ifConnected = False
    socketObj = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    result = socketObj.connect_ex((ip, port))
    if result == 0:
        ifConnected = True
    else:
        ifConnected = False
    return ifConnected


# 判断URL最后的URI是否与输入的相同
def ifSameUri(uri, url):
    reFlag = False

    uri = uri.strip("/")
    urlUri = urllib.parse.urlparse(url).path.strip("/")

    reFlag = (uri==urlUri)
    return reFlag

# 传入一个URL，去除该URL参数的值，返回类似：http:domain:port/a1=&a2= 的字符串
def parseUrlWithoutArgsValue(url):
    reUrl = ""
    splitList = urllib.parse.urlsplit(url)
    nowArgs = splitList[3]
    # 去除参数值
    tmpArgsList = nowArgs.split("&")
    if len(tmpArgsList)==1 and tmpArgsList[0]=="":
        reUrl = url
    else:
        for argIndex,tmpArg in enumerate(tmpArgsList):
            tmpArgList = tmpArg.split("=")
            tmpList = tmpArgList[:1]+[""]
            tmpArgsList[argIndex]="=".join(tmpList)
        finalArgStr = "&".join(tmpArgsList)

        reUrl=urllib.parse.urlunsplit(tuple(list(splitList[:3])+[finalArgStr]+list(splitList[4:])))

    return reUrl


def getHtmlCrawlerScriptFileSuffixList():
    scriptFileSuffixList = []

    scriptFileSuffixList.append("js")

    return scriptFileSuffixList

def getHtmlCrawlerCssFileSuffixList():
    cssFileSuffixList = []

    cssFileSuffixList.append("css")

    return cssFileSuffixList


def getHtmlCrawlerSolveStaticFileSuffixList():
    solveFileSuffixList = []
    solveFileSuffixList += getHtmlCrawlerScriptFileSuffixList()
    solveFileSuffixList+=getHtmlCrawlerCssFileSuffixList()
    solveFileSuffixList.append("xml")
    return solveFileSuffixList


def getHtmlCrawlerNotSolveResourceFileSuffixList():
    notSolveFileSuffixList = []
    # 图片后缀
    notSolveFileSuffixList += [a.strip() for a in
                               "jpg, jpeg, png, gif, webp, bmp, tiff, svg, ico, heic, psd, raw".split(",")]
    # 视频后缀
    notSolveFileSuffixList += [a.strip() for a in
                               "mp4, avi, mov, mkv, webm, flv, wmv, mpeg, mpg, 3gp, m4v, rmvb, vob".split(",")]
    # 音频后缀
    notSolveFileSuffixList += [a.strip() for a in
                               "mp3, wav, aac, ogg, flac, wma, midi, m4a, aiff, amr, opus".split(",")]

    # 字体后缀
    notSolveFileSuffixList += [a.strip() for a in
                               "ttf, otf, woff, woff2, eot, ttc, dfont, svg, svgz, pfb, pfm, bdf, pcf, ufo, vfb".split(",")]

    return notSolveFileSuffixList

def getHtmlCrawlerNotCrawlerFileSuffixList():
    notSolveFileSuffixList = []
    # 压缩包后缀
    notSolveFileSuffixList += [a.strip() for a in
                               "zip, rar, 7z, tar, gz, bz2, xz, tar gz, tar bz2, tar xz, tgz, z, lz, lzma, lzh, arj, cab, iso, pkg, deb, rpm".split(
                                   ",")]

    # 二进制文件后缀
    notSolveFileSuffixList += [a.strip() for a in
                               "exe, dll, so, dylib, a, lib, bin, img, iso, dmg, sys, o, class, jar, pyc, mo, dat, db, rom, hex".split(
                                   ",")]

    return notSolveFileSuffixList


def extractUrls(text):
    """
    从文本中提取完整URL（支持http、https、ftp、file协议）
    匹配内容包括：协议、域名/IP、端口、路径、查询参数
    """
    pattern = re.compile(r"""
                (?i)                # 忽略大小写
                (?:https?|ftp|file)(?::|%3A)(?:/|%2F|(?:\\u002F)){2}  # 协议
                (?:[^:@/]+(?::[^@/]*)?@)?  # 用户名密码
                (?:
                    (?:[a-zA-Z0-9-]+\.)+[a-zA-Z]{2,}  # 域名
                    | (?:\d{1,3}\.){3}\d{1,3}       # IPv4地址
                )
                (?::\d+)?           # 端口号
                (?:
                    (?:/|%2F|(?:\\u002F))[a-zA-Z0-9-._~%@#]+     # 路径及片段
                )*
                (?:/|%2F|(?:\\u002F))*         # 匹配结尾的/
                (?:
                    \?(?:[a-zA-Z0-9-._~%@&=+/]|(?:\\u002F))+
                )*     # 查询参数
            """, re.VERBOSE)

    reList = pattern.findall(text)

    return reList


def extractUris(text):
    """
    从文本中提取URI（/开头，直到特殊字符或空格结尾）
    """
    pattern = re.compile(r"""
                (?i)                # 忽略大小写
                (?<![<])
                [.]*                # 匹配相对路径
                (?:
                    (?:/|%2F|(?:\\u002F))[a-zA-Z0-9-._~%@#]+     # 路径及片段
                )+
                (?:/|%2F|(?:\\u002F))*         # 匹配结尾的/
                (?:
                    \?(?:[a-zA-Z0-9-\._~%@&=+/]|(?:\\u002F))+
                )*     # 查询参数
            """, re.VERBOSE)

    reList = pattern.findall(text)

    return reList

def extractSpecilUris(text):
    """
    从文本中提取特殊URI（被指定字符包裹，不以/开头，但符合URI格式），包裹字符包括：
    1、引号（单双引号）
    """
    uriPattern = r"""
                (?:
                    [a-zA-Z0-9-\._~%@#]+(?:/|%2F|(?:\\u002F))     # 路径及片段
                )
                (?:
                    [a-zA-Z0-9-\._~%@#]+(?:/|%2F|(?:\\u002F))*     # 路径及片段
                )+
                (?:
                    \?(?:[a-zA-Z0-9-\._~%@&=+/]|(?:\\u002F))+
                )*     # 查询参数
    """


    startCharsList = []
    startCharsList.append({"chars":['"','"'],"stripChars":['"']})
    startCharsList.append({"chars":["'","'"],"stripChars":["'"]})

    patternList = []

    for tmpStartCharsDict in startCharsList:
        pattern = re.compile(r"(?i)"+tmpStartCharsDict["chars"][0]+uriPattern+tmpStartCharsDict["chars"][1], re.VERBOSE)
        patternList.append({"pattern":pattern,"stripChars":tmpStartCharsDict["stripChars"]})

    # 匹配所有被指定字符包裹的字符串
    firstSearchList = []

    for tmpPatternDict in patternList:
        tmpMatchStrList = tmpPatternDict["pattern"].findall(text)
        tmpStripChars = tmpPatternDict["stripChars"]
        for tmpStripChar in tmpStripChars:
            tmpMatchStrList = [s.strip(tmpStripChar) for s in tmpMatchStrList]
        firstSearchList += tmpMatchStrList
    return firstSearchList

def solveExtractedUrls(urlList):
    # 对提取到的结果进行处理
    for tmpIndex, tmpUrl in enumerate(urlList):
        # 将\u002F字符串替换为/
        partern = r"\\u002f"
        urlList[tmpIndex] = re.sub(partern, "/", urlList[tmpIndex],flags=re.I)

        # 进行一次URL解码
        urlList[tmpIndex] = urllib.parse.unquote(urlList[tmpIndex])

        # 判断url是否包含协议头
        tmpSplitedObj = urllib.parse.urlsplit(tmpUrl)
        if tmpSplitedObj.netloc!="" and tmpSplitedObj.scheme == "":
            urlList[tmpIndex] = urllib.parse.urlunsplit(tuple(["http"] + list(tmpSplitedObj[1:])))

    return urlList

def getAllUseDomainList(crawlerUrl,fileUrl,startUrl,extraUrlArr):
    reDomainList = []
    reDomainList.append(getUrlOnlyDomain(crawlerUrl))
    reDomainList.append(getUrlOnlyDomain(fileUrl))
    reDomainList.append(getUrlOnlyDomain(startUrl))
    for tmpExtraUrl in extraUrlArr:
        reDomainList.append(getUrlOnlyDomain(tmpExtraUrl))

    reDomainList = list(set(reDomainList))
    return reDomainList

def calcResponseHash(responseData, algorithm: str = 'md5') -> str:
    # 获取响应数据的bytes类型
    if type(responseData) == bytes:
        responseData = responseData
    elif type(responseData) == str:
        responseData = responseData.encode("utf-8")
    else:
        responseData = str(responseData).encode("utf-8")
    """计算数据的哈希值，默认使用MD5"""
    hasher = hashlib.new(algorithm)
    hasher.update(responseData)
    return hasher.hexdigest()